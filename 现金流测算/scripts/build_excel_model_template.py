"""
构建子洲县10万千瓦风电项目偿债现金流测算 Excel 模型
带完整 Excel 计算公式、颜色编码、专业格式

═══════════════════════════════════════════════════
  公式完整性原则（CRITICAL — 禁止违反）
═══════════════════════════════════════════════════
  ① 只有「假设参数」Sheet 中可直接写数字（蓝色字体）。
  ② 其余所有 Sheet（CFADS / 债务排期 / DSCR）的计算单元格
     必须写 Excel 公式字符串（以 "=" 开头），严禁直接写数值。
  ③ 所有公式通过跨 Sheet 引用「假设参数」中的数字，
     修改任何一个假设后全表自动重算——这是财务模型的核心价值。
  ④ 宽限期"还本=0"也需用公式 "=0" 写入，保持形式统一。
═══════════════════════════════════════════════════
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

wb = Workbook()

# ───────────────────────────────────────────────
# 颜色常量
# ───────────────────────────────────────────────
C_INPUT_FONT   = "0000FF"   # 蓝：硬编码输入
C_FORMULA_FONT = "000000"   # 黑：公式单元格
C_HEADER_BG    = "1F4E79"   # 深蓝表头背景
C_HEADER_FONT  = "FFFFFF"   # 白色表头字体
C_SECTION_BG   = "D6E4F0"   # 浅蓝分区背景
C_WARN_BG      = "FFF2CC"   # 黄：关键假设
C_RED_BG       = "FFE0E0"   # 红：预警区
C_GREEN_BG     = "E2EFDA"   # 绿：充裕区
C_ALT_ROW      = "F5F9FF"   # 间隔行
C_BORDER       = "8EA9C1"   # 边框色

def hdr_fill(color=C_HEADER_BG):
    return PatternFill("solid", fgColor=color)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(bold=True, color=C_HEADER_FONT, size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)

def input_font(size=9):
    return Font(name="Arial", bold=False, color=C_INPUT_FONT, size=size)

def formula_font(size=9, bold=False):
    return Font(name="Arial", bold=bold, color=C_FORMULA_FONT, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center")

FMT_WAN  = '#,##0.00'          # 万元两位小数
FMT_PCT  = '0.00%'             # 百分比
FMT_NUM  = '#,##0.00'
FMT_DSCR = '0.00"x"'          # DSCR  显示 3.56x
FMT_ZERO = '#,##0.00;(#,##0.00);"-"'  # 零显示为 -

def set_cell(ws, row, col, value, font=None, fill=None, align=None,
             border=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt
    return cell

# ═══════════════════════════════════════════════════════════════
# Sheet 1 ── 假设参数区  (Assumptions)
# ═══════════════════════════════════════════════════════════════
ws_a = wb.active
ws_a.title = "假设参数"

# 调整列宽
ws_a.column_dimensions['A'].width = 28
ws_a.column_dimensions['B'].width = 16
ws_a.column_dimensions['C'].width = 22
ws_a.column_dimensions['D'].width = 16
ws_a.column_dimensions['E'].width = 22
ws_a.column_dimensions['F'].width = 16

def title_row(ws, row, text, color=C_HEADER_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill  = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    ws.row_dimensions[row].height = 18

def section_hdr(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Arial", bold=True, color="1F4E79", size=9)
    c.fill  = PatternFill("solid", fgColor=C_SECTION_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = thin_border()
    ws.row_dimensions[row].height = 15

def param_row(ws, row, label, val1, note1="", val2=None, note2="", is_formula=False):
    fill_ = PatternFill("solid", fgColor="FFFFFF") if row % 2 == 0 else PatternFill("solid", fgColor=C_ALT_ROW)
    f_val = formula_font() if is_formula else input_font()
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border()
        ws.cell(row=row, column=col).fill   = fill_
    ws.cell(row=row, column=1, value=label).font      = formula_font()
    ws.cell(row=row, column=1).alignment = left()
    set_cell(ws, row, 2, val1, font=f_val, align=right())
    set_cell(ws, row, 3, note1, font=formula_font(), align=left())
    if val2 is not None:
        set_cell(ws, row, 4, label, font=formula_font(), align=left())
        set_cell(ws, row, 5, val2, font=f_val, align=right())
        set_cell(ws, row, 6, note2, font=formula_font(), align=left())

# ── 主标题 ──
title_row(ws_a, 1, "子洲县10万千瓦风电项目  ──  偿债现金流测算模型  |  参数假设区")
ws_a.row_dimensions[1].height = 22

# ── 第一节：项目基本信息 ──
section_hdr(ws_a, 2, "一、项目基本信息")
rows_info = [
    ("项目名称",        "子洲县10万千瓦风电项目",   "（陕西省榆林市子洲县马岔镇）"),
    ("建设单位",        "中节能陕西风力发电有限公司", ""),
    ("装机容量 (MW)",   100,                         "16台×6,250kW"),
    ("年上网电量 (MWh)",243569.5,                    "满负荷年发电量"),
    ("年满负荷小时数",  2435.7,                      "h/年"),
    ("总动态投资 (万元)",57225.63,                   "含建设期利息"),
]
for i, (label, val, note) in enumerate(rows_info, start=3):
    param_row(ws_a, i, label, val, note)

# ── 第二节：融资参数（关键输入，黄底高亮）──
section_hdr(ws_a, 9, "二、融资结构参数（蓝色=可调输入）")
financing = [
    ("贷款本金 (万元)",      45780.0,  "总投资×80%",         0.0345,     "年化利率（LPR当前水平）"),
    ("资本金 (万元)",        "=B10-B11","=C11",             13,         "贷款期限（年）"),
    ("宽限期（年）",         1,        "建设期仅付息不还本",  12,         "还款期=贷款期-宽限期"),
    ("半年还本额 (万元)",    1907.50,  "每半年等额归还",      3815.0,     "年度等额还本（×2）"),
    ("还本期数（期）",        24,       "=还款期×2",          "=B13*2",  "验证：24期×1907.5=45780"),
]
# 手动写这几行（有公式）
r = 10
# 行10：贷款本金
param_row(ws_a, r,   "贷款本金 (万元)",   45780.0,  "总投资的80%，单位：万元",  0.0345, "年化利率（LPR当前水平 3.45%）")
ws_a.cell(r,2).number_format = FMT_WAN
ws_a.cell(r,5).number_format = FMT_PCT
ws_a.cell(r,4).value = "年化利率"
ws_a.cell(r,4).font  = formula_font()
ws_a.cell(r,4).alignment = left()

r = 11
param_row(ws_a, r,   "资本金 (万元)",   "=B10*(1-E10)", "=总投资×20%，公式联动",  13, "贷款期限（年，含宽限期）")
ws_a.cell(r,2).font  = formula_font()
ws_a.cell(r,2).number_format = FMT_WAN
ws_a.cell(r,4).value = "贷款期限（年）"
ws_a.cell(r,4).font  = formula_font()
ws_a.cell(r,4).alignment = left()
ws_a.cell(r,5).font  = input_font()

r = 12
param_row(ws_a, r,   "宽限期（年）",   1,  "建设期仅付息，不归还本金",  "=E11-B12", "还款期（年）= 贷款期 - 宽限期")
ws_a.cell(r,5).font  = formula_font()

r = 13
param_row(ws_a, r,   "半年还本额 (万元)", 1907.50, "每半年等额归还本金", "=B13*2", "年度等额还本（万元）=半年×2")
ws_a.cell(r,5).font  = formula_font()
ws_a.cell(r,2).number_format = FMT_WAN
ws_a.cell(r,5).number_format = FMT_WAN

r = 14
param_row(ws_a, r,   "还款期数（期）",  "=E12*2", "=还款年×2，共24期",  "=B13*E12*2", "验证：应=贷款本金（45,780万）")
ws_a.cell(r,2).font  = formula_font()
ws_a.cell(r,5).font  = formula_font()
ws_a.cell(r,5).number_format = FMT_WAN
# 条件标注
ws_a.cell(r,5).fill = PatternFill("solid", fgColor=C_WARN_BG)

# ── 第三节：发电参数 ──
section_hdr(ws_a, 15, "三、发电收入参数（蓝色=可调输入）")
r = 16
param_row(ws_a, r,  "含税上网电价（元/kWh）", 0.265,  "国网收购价，含增值税",  0.09, "增值税税率")
ws_a.cell(r,2).number_format = '0.000'
ws_a.cell(r,5).number_format = FMT_PCT
ws_a.cell(r,4).value = "增值税率"
ws_a.cell(r,4).font  = formula_font()
ws_a.cell(r,4).alignment = left()

r = 17
param_row(ws_a, r,  "不含税电价（元/kWh）",  "=B16/(1+E16)", "=含税÷(1+增值税率)", "", "")
ws_a.cell(r,2).font = formula_font()
ws_a.cell(r,2).number_format = '0.0000'

r = 18
param_row(ws_a, r,  "年达产率", 0.96, "含弃风+线损+停机折减", "", "")
ws_a.cell(r,2).number_format = FMT_PCT

r = 19
param_row(ws_a, r, "年上网电量（MWh）", 243569.5, "满负荷年度上网电量", "", "")
ws_a.cell(r,2).number_format = '#,##0.0'

r = 20
param_row(ws_a, r, "年核心收入（万元/年）",
          "=B19*B18*B16*1000/10000",
          "=上网电量×达产率×含税电价×1000÷10000（MWh→kWh→万元）",
          "", "")
ws_a.cell(r,2).font = formula_font()
ws_a.cell(r,2).number_format = FMT_WAN
ws_a.cell(r,2).fill = PatternFill("solid", fgColor=C_WARN_BG)

# ── 第四节：OPEX参数 ──
section_hdr(ws_a, 21, "四、运营成本参数（OPEX，万元/年）")
r = 22
for label, val1, note1, val2, note2 in [
    ("维修费（保内1-6年）",     100,  "厂商维保合同",         300,  "保外7-8年（万元）"),
    ("维修费（保外9-10年）",    400,  "万元/年",              500,  "保外11-12年（万元）"),
    ("其他不可预计（保内）",     100,  "1-6年，万元/年",       200,  "7-8年（万元）"),
    ("其他不可预计（保外9-10）", 250,  "万元/年",              300,  "11-12年（万元）"),
    ("土地租金",                 18,   "固定，万元/年",         200,  "人工成本（5人×40万/人）"),
    ("保险费",                   140,  "总资产×0.25%约140万",  0,    "（含在OPEX合计中）"),
]:
    param_row(ws_a, r, label, val1, note1, val2, note2)
    r += 1

# ── 第五节：税率 ──
section_hdr(ws_a, 28, "五、税率假设（三免三减半+西部大开发）")
r = 29
for label, val, note in [
    ("第1-3年综合税率",  0.085, "免所得税，仅增值税及附加约8.5%"),
    ("第4-6年综合税率",  0.105, "所得税7.5%（减半）+增值税附加"),
    ("第7年起综合税率",  0.125, "所得税15%（西部大开发优惠）+增值税附加"),
    ("维持性CAPEX（保内1-6年）",    0,   "万元/年，厂商包修无额外大修"),
    ("维持性CAPEX（保外7-12年）",   500, "万元/年，叶片/齿轮箱大检修准备金"),
]:
    param_row(ws_a, r, label, val, note)
    ws_a.cell(r,2).number_format = FMT_PCT if "税率" in label else FMT_WAN
    r += 1

# ── 第六节：DSCR 警戒参数 ──
section_hdr(ws_a, 34, "六、DSCR 合规参数")
param_row(ws_a, 35, "DSCR 警戒线（最低值）",  1.20, "银行贷款合同约定最低值")
param_row(ws_a, 36, "DSCR 舒适线",             1.50, "建议维持水平")
ws_a.cell(35,2).number_format = '0.00"x"'
ws_a.cell(36,2).number_format = '0.00"x"'

# 冻结首行+首列
ws_a.freeze_panes = "B2"


# ═══════════════════════════════════════════════════════════════
# Sheet 2 ── CFADS 现金流瀑布  (CF Waterfall)
# ═══════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet("CFADS现金流瀑布")
ws_cf.column_dimensions['A'].width = 22
for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N']:
    ws_cf.column_dimensions[col].width = 13

# 年份列（B~M = 运营第1~12年）
YEAR_START_COL = 2   # 列B
NUM_YEARS = 12

def cf_hdr(ws, row, text, merge_end_col=14, color=C_HEADER_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill  = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

cf_hdr(ws_cf, 1, "子洲县10万千瓦风电项目  ──  CFADS 现金流瀑布（单位：万元）")

# ── 列标题 ──
hdr_r = 2
ws_cf.cell(hdr_r, 1, "科目").font  = hdr_font(size=9)
ws_cf.cell(hdr_r, 1).fill         = hdr_fill()
ws_cf.cell(hdr_r, 1).alignment    = center()
ws_cf.cell(hdr_r, 1).border       = thin_border()
for y in range(NUM_YEARS):
    col = YEAR_START_COL + y
    c = ws_cf.cell(hdr_r, col, value=f"运营第{y+1:02d}年")
    c.font  = hdr_font(size=9)
    c.fill  = hdr_fill()
    c.alignment = center()
    c.border = thin_border()
# 合计列
ws_cf.cell(hdr_r, 14, "合计").font  = hdr_font()
ws_cf.cell(hdr_r, 14).fill         = hdr_fill()
ws_cf.cell(hdr_r, 14).alignment    = center()
ws_cf.cell(hdr_r, 14).border       = thin_border()

# ── 数据行定义 ──
# 参考：假设参数 sheet 名称="假设参数"
A = "假设参数"   # 简写

# 每行：(标签, [12个公式或值], 格式, 是否加粗, 是否为分隔行)
# 使用命名单元格引用假设参数，格式：假设参数!B20 等
def yr_ref(abs_row):
    """返回假设参数中某行B列的绝对引用（所有年相同的全局参数）"""
    return f"'{A}'!$B${abs_row}"

# 含税收入（每年相同）
inc_formula   = [f"='{A}'!$B$20" for _ in range(NUM_YEARS)]  # 年核心收入

# OPEX：按年段差异化
def opex_formula(y):
    # y: 0-indexed
    # 保内1-6：100+100+18+200+140=558，保外7-8：858，9-10：1008，11-12：1158
    # 引用假设参数表的参数单元格重新计算
    if y < 6:
        return (f"='{A}'!$B$22+'{A}'!$B$24"
                f"+'{A}'!$B$26+'{A}'!$B$27+'{A}'!$B$28")
    elif y < 8:
        return (f"='{A}'!$E$22+'{A}'!$E$24"
                f"+'{A}'!$B$26+'{A}'!$B$27+'{A}'!$B$28")
    elif y < 10:
        return (f"='{A}'!$B$23+'{A}'!$B$25"
                f"+'{A}'!$B$26+'{A}'!$B$27+'{A}'!$B$28")
    else:
        return (f"='{A}'!$E$23+'{A}'!$E$25"
                f"+'{A}'!$B$26+'{A}'!$B$27+'{A}'!$B$28")

def tax_formula(y):
    # 税率段落: 1-3年→B29, 4-6年→B30, 7-12年→B31
    if y < 3:
        rate_ref = f"'{A}'!$B$29"
    elif y < 6:
        rate_ref = f"'{A}'!$B$30"
    else:
        rate_ref = f"'{A}'!$B$31"
    income_col = get_column_letter(YEAR_START_COL + y)
    return f"={income_col}4*{rate_ref}"

def capex_formula(y):
    if y < 6:
        return f"='{A}'!$B$32"   # 0
    else:
        return f"='{A}'!$E$33"   # 500

# 表格行数据：
# 行3：空白分隔
# 行4：核心收入
# 行5：(-) OPEX
# 行6：(-) 税费
# 行7：(-) 维持性CAPEX
# 行8：分隔线
# 行9：CFADS合计

ROW_INC   = 4
ROW_OPEX  = 5
ROW_TAX   = 6
ROW_CAPEX = 7
ROW_CFADS = 8

cf_rows = [
    (ROW_INC,   "（+）核心收入（含税）", inc_formula,                   FMT_WAN, False),
    (ROW_OPEX,  "（-）付现OPEX（万元）", [opex_formula(y) for y in range(NUM_YEARS)], FMT_WAN, False),
    (ROW_TAX,   "（-）综合税费（万元）", [tax_formula(y) for y in range(NUM_YEARS)],  FMT_WAN, False),
    (ROW_CAPEX, "（-）维持性CAPEX",      [capex_formula(y) for y in range(NUM_YEARS)],FMT_WAN, False),
    (ROW_CFADS, "★ CFADS（可用于还本付息）",
                [f"={get_column_letter(YEAR_START_COL+y)}{ROW_INC}"
                 f"-{get_column_letter(YEAR_START_COL+y)}{ROW_OPEX}"
                 f"-{get_column_letter(YEAR_START_COL+y)}{ROW_TAX}"
                 f"-{get_column_letter(YEAR_START_COL+y)}{ROW_CAPEX}"
                 for y in range(NUM_YEARS)],
                FMT_WAN, True),
]

for row_num, label, formulas, fmt, bold in cf_rows:
    fill_color = C_WARN_BG if bold else ("FFFFFF" if row_num % 2 == 0 else C_ALT_ROW)
    c = ws_cf.cell(row_num, 1, value=label)
    c.font  = Font(name="Arial", bold=bold, size=9)
    c.fill  = PatternFill("solid", fgColor=fill_color if not bold else "D9F0D3")
    c.alignment = left()
    c.border = thin_border()
    for y, formula in enumerate(formulas):
        col = YEAR_START_COL + y
        cell = ws_cf.cell(row_num, col, value=formula)
        cell.font   = formula_font(bold=bold)
        cell.fill   = PatternFill("solid", fgColor=fill_color if not bold else "D9F0D3")
        cell.number_format = fmt
        cell.alignment = right()
        cell.border = thin_border()
    # 合计列
    sum_range = f"{get_column_letter(YEAR_START_COL)}{row_num}:{get_column_letter(YEAR_START_COL+NUM_YEARS-1)}{row_num}"
    c_sum = ws_cf.cell(row_num, 14, value=f"=SUM({sum_range})")
    c_sum.font = formula_font(bold=bold)
    c_sum.fill = PatternFill("solid", fgColor=fill_color if not bold else "D9F0D3")
    c_sum.number_format = fmt
    c_sum.alignment = right()
    c_sum.border = thin_border()

# 税率注释行
ws_cf.cell(9, 1, "适用综合税率").font = formula_font()
ws_cf.cell(9, 1).alignment = left()
ws_cf.cell(9, 1).border = thin_border()
for y in range(NUM_YEARS):
    col = YEAR_START_COL + y
    if y < 3:
        rate_ref = f"='{A}'!$B$29"
    elif y < 6:
        rate_ref = f"='{A}'!$B$30"
    else:
        rate_ref = f"='{A}'!$B$31"
    c = ws_cf.cell(9, col, value=rate_ref)
    c.font = formula_font()
    c.number_format = FMT_PCT
    c.alignment = center()
    c.border = thin_border()
ws_cf.cell(9, 14, "—").alignment = center()
ws_cf.cell(9, 14).border = thin_border()

ws_cf.freeze_panes = "B3"


# ═══════════════════════════════════════════════════════════════
# Sheet 3 ── 债务偿还排期  (Debt Schedule)
# ═══════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("债务偿还排期")
ws_d.column_dimensions['A'].width = 8
for col_letter in ['B','C','D','E','F','G','H']:
    ws_d.column_dimensions[col_letter].width = 15

# 总期限 13 年（含1年宽限期）
TOTAL_YEARS = 13
GRACE_YEARS = 1

def ds_hdr(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

ds_hdr(ws_d, 1, "债务偿还排期  ──  含宽限期利息（单位：万元）")

# 列标题
headers_d = ["年份", "类型", "期初余额", "当期还本", "当期利息", "偿债合计", "期末余额", "备注"]
for c_idx, h in enumerate(headers_d, start=1):
    cell = ws_d.cell(2, c_idx, value=h)
    cell.font  = hdr_font()
    cell.fill  = hdr_fill()
    cell.alignment = center()
    cell.border = thin_border()
ws_d.row_dimensions[2].height = 16

# 数据行 (13行)：Y01~Y13
# 参数引用假设参数 sheet
PA = f"'{A}'"
# 宽限期1年（Y01），还款期2~13年（Y02~Y13）
for yr in range(1, TOTAL_YEARS + 1):
    row = yr + 2
    is_grace = yr <= GRACE_YEARS
    fill_color = "FFF8E7" if is_grace else ("FFFFFF" if yr % 2 == 0 else C_ALT_ROW)
    fill_ = PatternFill("solid", fgColor=fill_color)

    # A: 年份
    c = ws_d.cell(row, 1, value=f"Y{yr:02d}")
    c.font = formula_font(bold=is_grace)
    c.fill = fill_
    c.alignment = center()
    c.border = thin_border()

    # B: 类型
    type_val = "宽限期（仅付息）" if is_grace else "还款期（等额本金）"
    c = ws_d.cell(row, 2, value=type_val)
    c.font = formula_font(bold=is_grace)
    c.fill = fill_
    c.alignment = center()
    c.border = thin_border()

    # C: 期初余额
    if yr == 1:
        opening_formula = f"={PA}!$B$10"  # 贷款本金
    else:
        opening_formula = f"=G{row-1}"     # 上期期末余额
    c = ws_d.cell(row, 3, value=opening_formula)
    c.font = formula_font()
    c.fill = fill_
    c.number_format = FMT_WAN
    c.alignment = right()
    c.border = thin_border()

    # D: 当期还本
    if is_grace:
        # 宽限期：还本=0，用公式形式写入（保持公式一致性，勿直接写数字0）
        repay_formula = "=0"
        c = ws_d.cell(row, 4, value=repay_formula)
        c.font = formula_font()
    else:
        repay_formula = f"={PA}!$E$13"   # 年度等额还本 3,815
        c = ws_d.cell(row, 4, value=repay_formula)
        c.font = formula_font()
    c.fill = fill_
    c.number_format = FMT_WAN
    c.alignment = right()
    c.border = thin_border()

    # E: 当期利息 = 期初余额 × 年利率
    interest_formula = f"=C{row}*{PA}!$E$10"
    c = ws_d.cell(row, 5, value=interest_formula)
    c.font = formula_font()
    c.fill = fill_
    c.number_format = FMT_WAN
    c.alignment = right()
    c.border = thin_border()

    # F: 偿债合计 = 还本 + 利息
    ds_formula = f"=D{row}+E{row}"
    c = ws_d.cell(row, 6, value=ds_formula)
    c.font = formula_font(bold=True)
    c.fill = fill_
    c.number_format = FMT_WAN
    c.alignment = right()
    c.border = thin_border()

    # G: 期末余额 = 期初 - 还本
    closing_formula = f"=C{row}-D{row}"
    c = ws_d.cell(row, 7, value=closing_formula)
    c.font = formula_font()
    c.fill = fill_
    c.number_format = FMT_WAN
    c.alignment = right()
    c.border = thin_border()

    # H: 备注
    note = "建设期利息资本化" if is_grace else ""
    c = ws_d.cell(row, 8, value=note)
    c.font = formula_font()
    c.fill = fill_
    c.alignment = left()
    c.border = thin_border()

# 合计行
sum_row = TOTAL_YEARS + 3
ws_d.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=2)
c = ws_d.cell(sum_row, 1, value="合计")
c.font  = hdr_font()
c.fill  = hdr_fill()
c.alignment = center()
for r_col in range(1, 9):
    ws_d.cell(sum_row, r_col).border = thin_border()
    ws_d.cell(sum_row, r_col).fill   = hdr_fill()
for dc, fcol_start, sum_col in [(4,"D","D"),(5,"E","E"),(6,"F","F")]:
    sum_f = f"=SUM({sum_col}3:{sum_col}{TOTAL_YEARS+2})"
    c = ws_d.cell(sum_row, dc, value=sum_f)
    c.font = hdr_font()
    c.number_format = FMT_WAN
    c.alignment = right()
ws_d.cell(sum_row, 7, value="─").alignment = center()
ws_d.cell(sum_row, 7).font = hdr_font()

ws_d.freeze_panes = "C3"


# ═══════════════════════════════════════════════════════════════
# Sheet 4 ── DSCR 综合测算  (DSCR Dashboard)
# ═══════════════════════════════════════════════════════════════
ws_ds = wb.create_sheet("DSCR综合测算")
ws_ds.column_dimensions['A'].width = 10
for col_letter in ['B','C','D','E','F','G','H','I','J']:
    ws_ds.column_dimensions[col_letter].width = 14

def dscr_hdr(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

dscr_hdr(ws_ds, 1, "DSCR 综合测算表  ──  偿债备付率分析（警戒线 ≥ 1.20x）")

# 列标题
headers_ds = ["运营年份","贷款年份","CFADS（万元）","期初余额","当期还本",
              "当期付息","偿债合计","期末余额","DSCR","状态"]
for c_idx, h in enumerate(headers_ds, start=1):
    cell = ws_ds.cell(2, c_idx, value=h)
    cell.font  = hdr_font()
    cell.fill  = hdr_fill()
    cell.alignment = center()
    cell.border = thin_border()
ws_ds.row_dimensions[2].height = 16

# 数据行 12 行（运营第1~12年 = 贷款 Y02~Y13）
CF = "CFADS现金流瀑布"
DS = "债务偿还排期"

for op_yr in range(1, NUM_YEARS + 1):
    row = op_yr + 2
    loan_yr = op_yr + GRACE_YEARS    # 运营第1年=贷款Y02, 对应debt sheet行 op_yr+1+2=op_yr+3

    # 债务排期中对应行
    # 债务排期：数据从第3行起，Y01=row3, Y02=row4... Y13=row15
    # 运营第1年 → Y02 → 债务排期第4行
    ds_data_row = loan_yr + 2   # Y02→row4, Y03→row5 ...

    fill_color = "FFFFFF" if op_yr % 2 == 0 else C_ALT_ROW

    # A: 运营年份
    c = ws_ds.cell(row, 1, value=f"第{op_yr:02d}年")
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.alignment = center(); c.border = thin_border()

    # B: 贷款年份
    c = ws_ds.cell(row, 2, value=f"Y{loan_yr:02d}")
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.alignment = center(); c.border = thin_border()

    # C: CFADS（从CFADS sheet读取）
    # CFADS现金流瀑布第8行是CFADS合计行，列B~M对应运营第1~12年
    cfads_col = get_column_letter(YEAR_START_COL + op_yr - 1)
    cfads_f = f"='{CF}'!{cfads_col}8"
    c = ws_ds.cell(row, 3, value=cfads_f)
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = FMT_WAN; c.alignment = right(); c.border = thin_border()

    # D: 期初余额
    c = ws_ds.cell(row, 4, value=f"='{DS}'!C{ds_data_row}")
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = FMT_WAN; c.alignment = right(); c.border = thin_border()

    # E: 当期还本
    c = ws_ds.cell(row, 5, value=f"='{DS}'!D{ds_data_row}")
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = FMT_WAN; c.alignment = right(); c.border = thin_border()

    # F: 当期付息
    c = ws_ds.cell(row, 6, value=f"='{DS}'!E{ds_data_row}")
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = FMT_WAN; c.alignment = right(); c.border = thin_border()

    # G: 偿债合计
    c = ws_ds.cell(row, 7, value=f"=E{row}+F{row}")
    c.font = formula_font(bold=True); c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = FMT_WAN; c.alignment = right(); c.border = thin_border()

    # H: 期末余额
    c = ws_ds.cell(row, 8, value=f"='{DS}'!G{ds_data_row}")
    c.font = formula_font(); c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = FMT_WAN; c.alignment = right(); c.border = thin_border()

    # I: DSCR = CFADS / 偿债合计（防DIV/0）
    dscr_f = f"=IF(G{row}>0,C{row}/G{row},\"—\")"
    c = ws_ds.cell(row, 9, value=dscr_f)
    c.font = Font(name="Arial", bold=True, size=9, color=C_FORMULA_FONT)
    c.fill = PatternFill("solid", fgColor=fill_color)
    c.number_format = '0.00'; c.alignment = center(); c.border = thin_border()

    # J: 状态（公式判断）
    # 引用假设参数中的警戒线（B35）
    status_f = (f"=IF(I{row}=\"—\",\"─\","
                f"IF(I{row}>={PA}!$B$36,\"✅ 充裕\","
                f"IF(I{row}>={PA}!$B$35,\"🟡 合格\",\"🔴 预警\")))")
    c = ws_ds.cell(row, 10, value=status_f)
    c.font = formula_font(bold=True)
    c.fill = PatternFill("solid", fgColor=fill_color)
    c.alignment = center(); c.border = thin_border()

# ── 统计汇总区 ──
sum_start = NUM_YEARS + 4
ws_ds.merge_cells(start_row=sum_start, start_column=1, end_row=sum_start, end_column=10)
c = ws_ds.cell(sum_start, 1, value="▌ 综合指标摘要")
c.font  = Font(name="Arial", bold=True, size=10, color="1F4E79")
c.fill  = PatternFill("solid", fgColor=C_SECTION_BG)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
c.border = thin_border()

stats = [
    ("DSCR 最低值",    f"=MIN(I3:I{NUM_YEARS+2})",  '0.00"x"'),
    ("DSCR 平均值",    f"=AVERAGE(I3:I{NUM_YEARS+2})",'0.00"x"'),
    ("DSCR 警戒线",    f"={PA}!$B$35",               '0.00"x"'),
    ("CFADS 合计（万）", f"=SUM(C3:C{NUM_YEARS+2})", FMT_WAN),
    ("偿债合计（万）",  f"=SUM(G3:G{NUM_YEARS+2})",  FMT_WAN),
    ("利息合计（万）",  f"=SUM(F3:F{NUM_YEARS+2})",  FMT_WAN),
    ("低于警戒年数",    f"=COUNTIF(I3:I{NUM_YEARS+2},\"<\"&{PA}!$B$35)", "0"),
]
for i, (label, formula, fmt) in enumerate(stats):
    stat_row = sum_start + 1 + i
    c1 = ws_ds.cell(stat_row, 1, value=label)
    c1.font = formula_font(bold=True)
    c1.alignment = left()
    c1.border = thin_border()
    ws_ds.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=2)

    c2 = ws_ds.cell(stat_row, 3, value=formula)
    c2.font = formula_font(bold=True)
    c2.number_format = fmt
    c2.alignment = center()
    c2.border = thin_border()
    for col in [1,2,3]:
        ws_ds.cell(stat_row, col).fill = PatternFill("solid", fgColor=C_ALT_ROW)

ws_ds.freeze_panes = "C3"


# ═══════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════
OUTPUT_PATH = r"D:\Workbuddy\wind_power_dscr_model.xlsx"
wb.save(OUTPUT_PATH)
print(f"Excel模型已保存：{OUTPUT_PATH}")
