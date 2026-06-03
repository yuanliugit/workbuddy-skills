#!/usr/bin/env python3
"""
excel_generator.py — 生成标准化财务数据底稿 Excel

用法：
  python excel_generator.py \
    --data D:\Workbuddy\企业披露信息抓取工作区\华天科技\01-财务报表\data.json \
    --output D:\Workbuddy\企业披露信息抓取工作区\华天科技\01-财务报表\财务数据底稿_20260521.xlsx \
    --company 华天科技 --years 2022 2023 2024
"""

import argparse
import json
import os
from datetime import datetime
from typing import Dict, Any, Optional, List


# ─────────────────────────────────────────────
# Excel 样式常量
# ─────────────────────────────────────────────
HEADER_BG = "2B5BAE"       # 深蓝色表头背景
HEADER_FG = "FFFFFF"       # 白色字体
SUBHEADER_BG = "D6E4F7"   # 浅蓝色副表头
SECTION_BG = "F0F5FF"     # 章节分隔行
NUMBER_FORMAT = '#,##0.00'  # 千分位两位小数


def safe_get(d: dict, *keys, default=None):
    """安全多级取值"""
    for key in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(key, {})
    return d if d != {} else default


def fmt_num(val) -> str:
    """格式化数值为万元字符串"""
    if val is None:
        return "-"
    try:
        return f"{float(val):,.2f}"
    except (TypeError, ValueError):
        return str(val)


def fmt_pct(val) -> str:
    """格式化为百分比"""
    if val is None:
        return "-"
    try:
        return f"{float(val) * 100:.2f}%"
    except (TypeError, ValueError):
        return str(val)


# ─────────────────────────────────────────────
# Sheet: 封面
# ─────────────────────────────────────────────
def write_cover_sheet(ws, company_name: str, years: list, warnings: list):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise

    ws.title = "封面"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50

    rows = [
        ("企业名称", company_name),
        ("数据年度", "、".join(years)),
        ("单位", "万元（人民币）"),
        ("生成日期", datetime.now().strftime("%Y年%m月%d日")),
        ("数据来源", "巨潮资讯网公开披露年度报告及审计报告"),
        ("", ""),
        ("说明", "本底稿数据来源于公开披露资料，已统一换算为万元。"),
    ]

    if warnings:
        rows.append(("数据异常", "；".join(warnings[:5])))

    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
        if k:
            ws.cell(i, 1).fill = header_fill
            ws.cell(i, 1).font = Font(bold=True, color=HEADER_FG)


# ─────────────────────────────────────────────
# Sheet: 三段式财务报表（合并/本部）
# ─────────────────────────────────────────────
BALANCE_SHEET_ITEMS = [
    ("资产负债表", None),              # 段落标题
    ("货币资金", "monetary_funds"),
    ("交易性金融资产", "trading_financial_assets"),
    ("应收票据", "notes_receivable"),
    ("应收账款", "accounts_receivable"),
    ("预付款项", "prepaid_expenses"),
    ("其他应收款", "other_receivables"),
    ("存货", "inventory"),
    ("流动资产合计", "current_assets"),
    ("固定资产", "fixed_assets"),
    ("在建工程", "construction_in_progress"),
    ("无形资产", "intangible_assets"),
    ("长期股权投资", "long_term_equity_investment"),
    ("非流动资产合计", "non_current_assets"),
    ("资产总计", "total_assets"),
    ("", None),
    ("短期借款", "short_term_loan"),
    ("应付票据", "notes_payable"),
    ("应付账款", "accounts_payable"),
    ("合同负债", "contract_liabilities"),
    ("一年内到期的非流动负债", "current_portion_lt_debt"),
    ("流动负债合计", "current_liabilities"),
    ("长期借款", "long_term_loan"),
    ("应付债券", "bonds_payable"),
    ("非流动负债合计", "non_current_liabilities"),
    ("负债合计", "total_liabilities"),
    ("", None),
    ("股本/实收资本", "share_capital"),
    ("资本公积", "capital_surplus"),
    ("盈余公积", "surplus_reserve"),
    ("未分配利润", "retained_earnings"),
    ("少数股东权益", "minority_interest"),
    ("所有者权益合计", "total_equity"),
]

INCOME_ITEMS = [
    ("利润表", None),
    ("营业收入", "revenue"),
    ("营业成本", "cost_of_revenue"),
    ("毛利润（计算）", "__gross_profit__"),
    ("税金及附加", "taxes_and_surcharges"),
    ("销售费用", "selling_expenses"),
    ("管理费用", "admin_expenses"),
    ("研发费用", "rd_expenses"),
    ("财务费用", "financial_expenses"),
    ("  其中：利息费用", "interest_expense"),
    ("资产减值损失", "asset_impairment_loss"),
    ("信用减值损失", "credit_impairment_loss"),
    ("投资收益", "investment_income"),
    ("营业利润", "operating_profit"),
    ("营业外收入", "non_operating_income"),
    ("营业外支出", "non_operating_expenses"),
    ("利润总额", "profit_before_tax"),
    ("所得税费用", "income_tax"),
    ("净利润", "net_profit"),
    ("归属于母公司股东净利润", "net_profit_attributable_to_parent"),
]

CASH_FLOW_ITEMS = [
    ("现金流量表", None),
    ("销售商品收到的现金", "cash_from_sales"),
    ("经营活动现金流入小计", "operating_cash_inflow"),
    ("经营活动现金流出小计", "operating_cash_outflow"),
    ("经营活动现金流量净额", "net_cash_from_operations"),
    ("资本性支出", "capex"),
    ("投资活动现金流量净额", "net_cash_from_investing"),
    ("取得借款收到的现金", "proceeds_from_borrowings"),
    ("偿还债务支付的现金", "repayment_of_debt"),
    ("筹资活动现金流量净额", "net_cash_from_financing"),
    ("现金净增加额", "net_increase_in_cash"),
    ("期末现金余额", "cash_at_end"),
]


def write_financial_sheet(ws, sheet_name: str, data_by_year: Dict[str, dict], years: list, scope: str):
    """写入三段式财务报表 Sheet"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise

    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    section_fill = PatternFill("solid", fgColor="4472C4")
    subtotal_fill = PatternFill("solid", fgColor=SUBHEADER_BG)

    # 设置列宽
    ws.column_dimensions["A"].width = 32
    for i, _ in enumerate(years, 2):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # 表头行
    ws.cell(1, 1, "科目").fill = header_fill
    ws.cell(1, 1).font = Font(bold=True, color=HEADER_FG)
    for i, yr in enumerate(years, 2):
        ws.cell(1, i, f"{yr}年（万元）").fill = header_fill
        ws.cell(1, i).font = Font(bold=True, color=HEADER_FG)
        ws.cell(1, i).alignment = Alignment(horizontal="center")

    row = 2
    all_items = BALANCE_SHEET_ITEMS + [("", None)] + INCOME_ITEMS + [("", None)] + CASH_FLOW_ITEMS

    for cn_name, key in all_items:
        # 空行
        if not cn_name and key is None:
            row += 1
            continue

        # 段落标题（蓝色背景）
        if key is None:
            cell = ws.cell(row, 1, cn_name)
            cell.fill = section_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            for i in range(2, len(years) + 2):
                ws.cell(row, i).fill = section_fill
            row += 1
            continue

        # 合计行加粗
        is_total = any(t in cn_name for t in ["合计", "净利润", "净额", "总计"])
        cell = ws.cell(row, 1, cn_name)
        if is_total:
            cell.font = Font(bold=True)
            for i in range(1, len(years) + 2):
                ws.cell(row, i).fill = PatternFill("solid", fgColor=SUBHEADER_BG)

        # 填充各年数据
        for i, yr in enumerate(years, 2):
            year_data = data_by_year.get(yr, {})
            fs_key = f"balance_sheet_{scope}" if cn_name in [x[0] for x in BALANCE_SHEET_ITEMS] else \
                     f"income_{scope}" if cn_name in [x[0] for x in INCOME_ITEMS] else \
                     f"cash_flow_{scope}"

            # 实际数据
            if key == "__gross_profit__":
                rev = year_data.get(f"income_{scope}", {}).get("revenue", 0) or 0
                cost = year_data.get(f"income_{scope}", {}).get("cost_of_revenue", 0) or 0
                val = rev - cost if rev else None
            else:
                val = None
                for prefix in [f"balance_sheet_{scope}", f"income_{scope}", f"cash_flow_{scope}"]:
                    v = year_data.get(prefix, {}).get(key)
                    if v is not None:
                        val = v
                        break

            c = ws.cell(row, i, val if val is not None else "")
            if val is not None and isinstance(val, (int, float)):
                c.number_format = NUMBER_FORMAT

        row += 1


# ─────────────────────────────────────────────
# Sheet: 关键财务指标
# ─────────────────────────────────────────────
def calc_indicators(bs: dict, is_: dict, cf: dict, prior_bs: dict = None) -> dict:
    """计算五类财务指标"""
    def safe_div(a, b, default=None):
        try:
            return a / b if b else default
        except Exception:
            return default

    ta = bs.get("total_assets", 0) or 0
    tl = bs.get("total_liabilities", 0) or 0
    ca = bs.get("current_assets", 0) or 0
    cl = bs.get("current_liabilities", 0) or 0
    inv = bs.get("inventory", 0) or 0
    pre = bs.get("prepaid_expenses", 0) or 0
    cash = bs.get("monetary_funds", 0) or 0
    equity = bs.get("total_equity", 0) or 0

    ibd = (bs.get("short_term_loan", 0) or 0) + \
          (bs.get("long_term_loan", 0) or 0) + \
          (bs.get("bonds_payable", 0) or 0) + \
          (bs.get("current_portion_lt_debt", 0) or 0)

    rev = is_.get("revenue", 0) or 0
    cost = is_.get("cost_of_revenue", 0) or 0
    np_ = is_.get("net_profit", 0) or 0
    pbt = is_.get("profit_before_tax", 0) or 0
    ie = is_.get("interest_expense", 0) or 0

    op_cf = cf.get("net_cash_from_operations", 0) or 0
    capex = abs(cf.get("capex", 0) or 0)

    prior_ta = (prior_bs.get("total_assets", ta) or ta) if prior_bs else ta
    prior_eq = (prior_bs.get("total_equity", equity) or equity) if prior_bs else equity
    avg_ta = (ta + prior_ta) / 2
    avg_eq = (equity + prior_eq) / 2

    ebit = pbt + ie

    return {
        "资产负债率": safe_div(tl, ta),
        "有息负债率": safe_div(ibd, ta),
        "流动比率": safe_div(ca, cl),
        "速动比率": safe_div(ca - inv - pre, cl),
        "利息保障倍数": safe_div(ebit, ie),
        "毛利率": safe_div(rev - cost, rev),
        "净利率": safe_div(np_, rev),
        "ROE（加权）": safe_div(np_, avg_eq),
        "ROA": safe_div(np_, avg_ta),
        "经营现金流/净利润": safe_div(op_cf, np_),
        "自由现金流（万元）": op_cf - capex,
        "收现比": safe_div(cf.get("cash_from_sales", 0) or 0, rev),
    }


def write_indicators_sheet(ws, data_by_year: Dict[str, dict], years: list):
    """写入关键财务指标 Sheet"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise

    ws.title = "关键财务指标"
    header_fill = PatternFill("solid", fgColor=HEADER_BG)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    for i, _ in enumerate(years, 3):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # 表头
    ws.cell(1, 1, "指标").fill = header_fill
    ws.cell(1, 1).font = Font(bold=True, color=HEADER_FG)
    ws.cell(1, 2, "类别").fill = header_fill
    ws.cell(1, 2).font = Font(bold=True, color=HEADER_FG)
    for i, yr in enumerate(years, 3):
        ws.cell(1, i, f"{yr}年").fill = header_fill
        ws.cell(1, i).font = Font(bold=True, color=HEADER_FG)

    CATEGORIES = {
        "资产负债率": "偿债能力", "有息负债率": "偿债能力", "流动比率": "偿债能力",
        "速动比率": "偿债能力", "利息保障倍数": "偿债能力",
        "毛利率": "盈利能力", "净利率": "盈利能力", "ROE（加权）": "盈利能力", "ROA": "盈利能力",
        "经营现金流/净利润": "现金流质量", "自由现金流（万元）": "现金流质量", "收现比": "现金流质量",
    }

    all_indicators_by_year = {}
    sorted_years = sorted(years)
    for i, yr in enumerate(sorted_years):
        bs = data_by_year.get(yr, {}).get("balance_sheet_consolidated", {})
        is_ = data_by_year.get(yr, {}).get("income_consolidated", {})
        cf = data_by_year.get(yr, {}).get("cash_flow_consolidated", {})
        prior_bs = data_by_year.get(sorted_years[i - 1], {}).get("balance_sheet_consolidated", {}) if i > 0 else None
        all_indicators_by_year[yr] = calc_indicators(bs, is_, cf, prior_bs)

    row = 2
    for metric, category in CATEGORIES.items():
        ws.cell(row, 1, metric)
        ws.cell(row, 2, category)
        for i, yr in enumerate(years, 3):
            val = all_indicators_by_year.get(yr, {}).get(metric)
            if val is None:
                ws.cell(row, i, "-")
            elif "万元" in metric:
                ws.cell(row, i, val).number_format = NUMBER_FORMAT
            else:
                ws.cell(row, i, val).number_format = "0.00%"
        row += 1


# ─────────────────────────────────────────────
# 主函数
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="生成财务数据底稿 Excel")
    parser.add_argument("--data", required=True, help="data.json 路径")
    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    parser.add_argument("--company", default="", help="公司名称")
    parser.add_argument("--years", nargs="+", help="年度列表，如 2022 2023 2024")
    args = parser.parse_args()

    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        print("安装依赖: pip install openpyxl")
        raise

    # 读取数据
    with open(args.data, "r", encoding="utf-8") as f:
        data_by_year = json.load(f)

    years = args.years if args.years else sorted(data_by_year.keys())
    company = args.company or "目标企业"

    print(f"=== 生成 Excel 底稿 ===")
    print(f"公司: {company}  年度: {years}")

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认Sheet

    # 收集所有警告
    all_warnings = []
    for yr in years:
        warnings = data_by_year.get(yr, {}).get("warnings", [])
        all_warnings.extend([f"{yr}年: {w}" for w in warnings])

    # Sheet 1: 封面
    ws_cover = wb.create_sheet("封面")
    write_cover_sheet(ws_cover, company, years, all_warnings)
    print("  ✅ 封面")

    # Sheet 2: 合并报表（三段式）
    ws_cons = wb.create_sheet("合并报表")
    write_financial_sheet(ws_cons, "合并报表", data_by_year, years, "consolidated")
    print("  ✅ 合并报表")

    # Sheet 3: 本部报表（三段式）
    ws_parent = wb.create_sheet("本部报表")
    write_financial_sheet(ws_parent, "本部报表", data_by_year, years, "parent")
    print("  ✅ 本部报表")

    # Sheet 4: 关键财务指标
    ws_ind = wb.create_sheet("关键财务指标")
    write_indicators_sheet(ws_ind, data_by_year, years)
    print("  ✅ 关键财务指标")

    # Sheet 5-6: 占位（业务板块、产能产销）
    wb.create_sheet("业务板块收入")
    wb.create_sheet("产能产销情况")

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)
    print(f"\n✅ Excel 已保存: {args.output}")


if __name__ == "__main__":
    main()
